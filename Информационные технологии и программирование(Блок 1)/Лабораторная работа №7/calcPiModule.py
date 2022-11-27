import math
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import openpyxl
def calcPi(e):
    pi = 0
    err = 1
    counter = int(0)
    try:
        workbook = openpyxl.load_workbook('tabResult.xlsx')
    except Exception:
        workbook = Workbook()
    sheet = workbook.active
    sheets = workbook.sheetnames
    sheet = workbook.active
    double = Side(border_style="double")
    while (abs(err) > e/4):
        err = (-1) ** counter * 1/(2 * counter + 1)
        pi += err
        cell = sheet.cell(row=counter + 1, column=2)
        cell.border = Border(left=double, bottom=double)
        cell.value = "Номер итерации:"
        cell = sheet.cell(row=counter + 1, column=3)
        cell.border = Border(bottom=double, right=double)
        cell.value = str(counter)
        cell = sheet.cell(row=counter + 1, column=4)
        cell.border = Border(bottom=double)
        cell.value = "Значение:"
        cell = sheet.cell(row=counter + 1, column=5)
        cell.border = Border(right=double, bottom=double)
        cell.value = str(pi * 4)
        counter += 1
        print("Номер итерации: " + str(counter) + " Вычисленное значения числа Пи: " + str(pi * 4))
    pi = pi * 4
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = (value+3)
    workbook.save('tabResult.xlsx')

    workbook = openpyxl.load_workbook('tabResult.xlsx')
    sheet = workbook.active
    for i in range(0, sheet.max_row):
        for col in sheet.iter_cols(1, sheet.max_column):
            print(col[i].value, end = "\t\t")
        print('')
    workbook.save('tabResult.xlsx')
    return pi